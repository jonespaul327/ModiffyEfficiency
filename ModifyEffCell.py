import openpyxl
import json
import os

def getListOfFiles(dirName):
    listOfFile = os.listdir(dirName)
    allFiles = list()
    for file in listOfFile:
        fullPath = os.path.join(dirName, file)
        if os.path.isdir(fullPath):
            allFiles = allFiles + getListOfFiles(fullPath)
        else:
            allFiles.append(fullPath)
    print(allFiles)
    return allFiles

def find_instrument_model_cell(currentSheet):
    for row in range(1, 50):
        for column in "ABCDEFGHIJKLMNOPQRSTUV":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if currentSheet[cell_name].value == instrumentModel:
                #print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                #print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
                print("the row is {0} and the column {1}" .format(cell_name[1], cell_name[0]))
                return cell_name

def find_instrument_sn_cell(instModelCell):
    snRow = str(int(instModelCell[1]) + 1)
    snCol = instModelCell[0]
    snCell = currentSheet[snCol + snRow]

    print('xxxxxxxxxxx')
    print(snRow)
    print(snCol)
    print(snCell.value)

    return snCell


def find_instrument_efficiency(instModelCell):
    effRow = str(int(instModelCell[1]) + 3)
    effCol = chr(ord(instModelCell[0]) + 2)
    effCell = currentSheet[effCol + effRow]

    print(effCell.value)
    print(effRow)
    print(effCol)

    return effCell

def modify_efficiency(instSNcell, instEfficiencyCell):
    for inst in instrumentsData:
        if inst['sn'] == instSNcell.value:
            instEfficiencyCell.value = inst['betaEfficiency']
            return inst['sn']


instrumentModel = '2360/43-93'
#instrumentId = '227413/PR295918'

filesWithNoSN = list()

files = getListOfFiles('surveys')

with open('package.json') as instruments_file:
    instrumentsData = json.load(instruments_file)

for file in files:
    theFile = openpyxl.load_workbook(file)
    allSheetNames = theFile.sheetnames

    print("All sheet names {} ".format(theFile.sheetnames))

    for x in allSheetNames:
        print("Current sheet name is {}" .format(x))
        currentSheet = theFile[x]
        instModelCell = find_instrument_model_cell(currentSheet)
        if instModelCell is None:
            continue
        instSNcell = find_instrument_sn_cell(instModelCell)
        instEfficiencyCell = find_instrument_efficiency(instModelCell)
        serialNumber = modify_efficiency(instSNcell, instEfficiencyCell)

        if serialNumber is None:
            filesWithNoSN.append(file)

        #print(currentSheet['L8'].value)

        # if currentSheet['L8'].value == instrumentId:
        #     currentSheet['N10'].value = 0.152

    theFile.close()
    theFile.save(file)

print("The files with no s/n are {}".format(filesWithNoSN))




"""
Purpose: This function is used to create a list of all files within a designated folder and then return the list. The 
            origin folder must be within the same folder as the program.
Parameter(s): The parameter is a folder name/folder path from the current directory of the program.
Return: Returns a list of all file names, including file path, within the folder as well as all subdirectories.

def getListOfFiles(dirName):
    listOfFile = os.listdir(dirName)
    allFiles = list()
    for file in listOfFile:
        fullPath = os.path.join(dirName, file)
        if os.path.isdir(fullPath):
            allFiles = allFiles + getListOfFiles(fullPath)
        else:
            allFiles.append(fullPath)
    print(allFiles)
    return allFiles
"""